import io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

FIELDS = [
    ('name',      'نام کسب‌وکار'),
    ('phone',     'تلفن'),
    ('address',   'آدرس'),
    ('city',      'شهر'),
    ('province',  'استان'),
    ('website',   'وبسایت'),
    ('rating',    'امتیاز'),
    ('category',  'دسته‌بندی'),
    ('latitude',  'عرض جغرافیایی'),
    ('longitude', 'طول جغرافیایی'),
    ('source',    'منبع'),
]

def make_xlsx(data, city='', category=''):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Results'
    ws.sheet_view.rightToLeft = True
    hfill = PatternFill('solid', fgColor='1F4E79')
    hfont = Font(bold=True, color='FFFFFF', size=11)
    for col, (_, label) in enumerate(FIELDS, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row_idx, item in enumerate(data, 2):
        for col, (key, _) in enumerate(FIELDS, 1):
            ws.cell(row=row_idx, column=col, value=str(item.get(key, '')))
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
